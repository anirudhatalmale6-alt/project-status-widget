import os
import openpyxl
from datetime import datetime


def read_projects(excel_path):
    """Read projects from Excel file. Returns list of dicts."""
    if not os.path.exists(excel_path):
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower().replace(' ', '_') if h else f'col_{i}'
               for i, h in enumerate(rows[0])]

    projects = []
    for row in rows[1:]:
        if not any(row):
            continue
        project = {}
        for i, val in enumerate(row):
            if i < len(headers):
                if isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d')
                project[headers[i]] = str(val) if val is not None else ''
        projects.append(project)

    wb.close()
    return projects


def search_projects(excel_path, query):
    """Search projects by name or claim number (case-insensitive partial match)."""
    projects = read_projects(excel_path)
    if not query:
        return []
    query = query.lower().strip()
    search_keys = [k for k in (projects[0].keys() if projects else [])
                   if 'name' in k or 'claim' in k or 'last' in k]
    if not search_keys:
        return [p for p in projects if any(query in str(v).lower() for v in p.values())]
    return [p for p in projects if any(query in str(p.get(k, '')).lower() for k in search_keys)]


def get_project_headers(excel_path):
    """Get column headers from Excel file."""
    if not os.path.exists(excel_path):
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()

    if not rows:
        return []
    return [str(h).strip() for h in rows[0] if h]
